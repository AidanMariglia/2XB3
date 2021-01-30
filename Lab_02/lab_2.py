import timeit
import random
from openpyxl import Workbook

def makelist(Length):
    list1 = []
    for i in range(Length):
        list1.append(i)
    return list1
    
def copy(List):
    List.copy()

def lookups():
    L = []
    filename = "lookups_rev_2.xlsx"
    workbook = Workbook()
    sheet = workbook.active


    for i in range(1000000):
        L.append(random.randrange(500))

    x = 0

    for i in range(1000000):
        start = timeit.default_timer()
        L[i]
        end = timeit.default_timer() - start
        sheet["A"+str(i+1)] = i
        sheet["B"+str(i+1)] = end

        #print(timeit.default_timer() - start)

    workbook.save(filename=filename)




def timetest(runs, Length):
    total = 0
    for _ in range(runs):
        list1 = makelist(Length)
        start = timeit.default_timer()
        copy(list1)
        end = timeit.default_timer()
        total += end - start
    return total/runs

def excel():
    Copyfile = "Copy.xlsx"
    workbook = Workbook()
    sheet = workbook.active

    for i in range(1000, 100000, 1000):
        total = 0
        sheet["A"+str(int(i/1000))] = i
        sheet["B"+str(int(i/1000))] = timetest(50, i)

    workbook.save(filename = Copyfile)

def append():
    filename = "append.xlsx"
    workbook = Workbook()
    sheet = workbook.active

    L = []

    for i in range(1000000):
        start = timeit.default_timer()
        L.append(1)
        end = timeit.default_timer() - start
        sheet["A" + str(i+1)] = i
        sheet["B" + str(i+1)] = end

    workbook.save(filename=filename)

append()

def multiAppend(n):
    L = []

    for i in range(n):
        L.append(i)


def multiAppendTime():
    filename = "experiments.xlsx"
    workbook = Workbook()
    sheet = workbook.active

    cell = 1

    for i in range(100000, 1000, 1000):
        start = timeit.default_timer()
        multiAppend(i)
        end = timeit.default_timer() - start

        sheet["A" + str(cell)] = i
        sheet["B" + str(cell)] = end

        cell += 1
        print(cell)

    workbook.save(filename=filename)
    

#excel()
#lookups()
multiAppendTime()